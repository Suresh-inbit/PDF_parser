import os
import pandas as pd
import google.generativeai as genai
from pathlib import Path
import json
from openpyxl import load_workbook

class ProposalExtractorGemini:
    def __init__(self, api_key, use_pro_model=True):
        """Initialize with Gemini API key
        
        Args:
            api_key: Your Gemini API key
            use_pro_model: If True, uses gemini-2.5-pro (highest accuracy, slower)
                          If False, uses gemini-2.5-flash (fast, good accuracy)
        """
        genai.configure(api_key=api_key)
        model_name = 'gemini-2.5-pro' if use_pro_model else 'gemini-2.5-flash'
        self.model = genai.GenerativeModel(model_name)
        print(f"Using model: {model_name}")
        
        # Define the extraction prompt
        self.extraction_prompt = """
You are analyzing a proposal PDF for quantum technologies lab setup. Extract the following information and respond ONLY with a JSON object.
THIS IS A VERY IMPORTANT TASK. FOLLOW THE INSTRUCTIONS CAREFULLY.
IMPORTANT OUTPUT REQUIREMENTS:
- For Yes/No fields: Return ONLY "Yes" or "No" or "Not Mentioned"(if you cannot find the information). In few case Yes/No may not be given explicitly instead might be in checkbox do this check if yes/no is not mentioned in the field.
- For text/qualitative fields: Extract the actual text/numbers from the document
- For ranking fields: Extract only the rank number (e.g., "45" not "Rank 45")

Extract these criteria:

COLUMN L: Central Government funded/aided institution - Yes/No only
COLUMN M: State Government funded/aided institution - Yes/No only  
COLUMN N: Other funding sources - Extract text describing other fundings that the institue is receiving(Not the funding for project), or write "None"
COLUMN O: Institution of Eminence by AICTE/UGC - Yes/No only
COLUMN P: NIRF 8th Edition rank number - Extract only the number (e.g., "45") or "Not Ranked"
COLUMN Q: QS World Ranking Asia 2024 rank number - Extract only the number (e.g., "120") or "Not Ranked"
COLUMN R: NBA accreditation for 30%+ courses (valid till April 2025) - Yes/No only
COLUMN S: NAAC score - Extract actual score number (e.g., "3.25") or "Not Available"
COLUMN T: Autonomous Status by UGC/AICTE(some propsal have not mentioned expicitly in that case check AICTE/UGC/Other Accreditation Status) - Return "UGC" or "AICTE" or "Both" or "None"
COLUMN U: More than 80% admission for last 5 years.(If not provided explicitly return Not Mentioned) - Yes/No only
COLUMN V: Two faculty members specified (1/3 time dedicated) - Yes/No only
COLUMN W: Lab space of at least 2000 sq.ft specified - Yes/No only
COLUMN X: Full-time lab technician specified - Yes/No only
COLUMN Y: Has the proposal specified that the institution has received Senate/Governing Body or Main Board of Institute approval for launching the UG Minor Programme in Quantum Technologies? - Yes/No only
COLUMN Z: Has the proposal submitted a formal written commitment indicating intent to start the programme upon obtaining the necessary approval?  - Yes/No only
COLUMN AA: Mention the page number of supporting document for COLUMN Y with comment on why yes/no(within 1 line)- Page Number(Comments) only
COLUMN AB: Mention the page number of supporting document for COLUMN Z with comment on why yes/no(within 1 line)- Page Number(Comments) only

Respond ONLY with this exact JSON format (no additional text):
{
    "col_l": "Yes or No",
    "col_m": "Yes or No",
    "col_n": "Text or None",
    "col_o": "Yes or No",
    "col_p": "Number or Not Ranked",
    "col_q": "Number or Not Ranked",
    "col_r": "Yes or No",
    "col_s": "Number or Not Available",
    "col_t": "UGC or AICTE or Both or None",
    "col_u": "Yes or No",
    "col_v": "Yes or No",
    "col_w": "Yes or No",
    "col_x": "Yes or No",
    "col_y": "Yes or No",
    "col_z": "Yes or No"
    "col_aa": "123(comment)"
    "col_ab": "123(comment)"

}
"""
    
    def extract_from_pdf(self, pdf_path, max_retries=2):
        """Extract criteria from PDF using Gemini with validation
        
        Args:
            pdf_path: Path to PDF file
            max_retries: Number of retries if extraction fails validation
        """
        import time
        from random import random

        print(f"  Uploading and analyzing PDF...")

        # Retry upload in case of transient network issues
        upload_attempts = 0
        pdf_file = None
        while upload_attempts <= max_retries:
            try:
                pdf_file = genai.upload_file(pdf_path)
                break
            except Exception as e:
                upload_attempts += 1
                wait = (2 ** upload_attempts) + random()
                print(f"  ⚠ Upload failed (attempt {upload_attempts}/{max_retries}). Error: {e}. Retrying in {wait:.1f}s...")
                time.sleep(wait)

        if pdf_file is None:
            print(f"  ⚠ Failed to upload PDF after {max_retries} attempts.")
            return None

        last_exception = None
        for attempt in range(max_retries + 1):
            try:
                response = self.model.generate_content([pdf_file, self.extraction_prompt])

                # Provide robust extraction of text portion
                response_text = getattr(response, 'text', '')
                if not response_text and hasattr(response, 'content'):
                    response_text = str(response.content)
                response_text = response_text.strip()

                # Remove markdown code blocks if present
                if response_text.startswith('```json'):
                    response_text = response_text.split('```json', 1)[1]
                if response_text.startswith('```'):
                    response_text = response_text.split('```', 1)[1]
                if response_text.endswith('```'):
                    response_text = response_text.rsplit('```', 1)[0]

                response_text = response_text.strip()

                # Parse JSON
                data = json.loads(response_text)

                # Clean up: delete the uploaded file from Gemini
                try:
                    pdf_file.delete()
                except Exception:
                    pass

                return data

            except json.JSONDecodeError as e:
                last_exception = e
                print(f"  ⚠ Error parsing JSON response: {str(e)}")
                snippet = (response_text[:200] + '...') if 'response_text' in locals() else ''
                print(f"  Raw response (truncated): {snippet}")
                if attempt < max_retries:
                    wait = (2 ** (attempt + 1)) + random()
                    print(f"  Retrying in {wait:.1f}s...")
                    time.sleep(wait)
                    continue
                else:
                    try:
                        pdf_file.delete()
                    except Exception:
                        pass
                    return None
            except Exception as e:
                last_exception = e
                # Inspect message to detect typical timeout codes
                err_msg = str(e)
                print(f"  ⚠ Error during analysis: {err_msg}")
                if attempt < max_retries:
                    wait = (2 ** (attempt + 1)) + random()
                    print(f"  Retrying (attempt {attempt + 2}/{max_retries + 1}) in {wait:.1f}s...")
                    time.sleep(wait)
                    continue
                else:
                    try:
                        pdf_file.delete()
                    except Exception:
                        pass
                    print(f"  ⚠ Final failure after {max_retries + 1} attempts: {last_exception}")
                    return None
    
    def update_excel_with_proposals(self, excel_path, pdf_folder, output_path=None):
        """Update an Excel file using TPN values from the sheet to find matching PDFs.

        The Excel is expected to have headers on row 5 (header=4 for pandas).
        """
        if output_path is None:
            output_path = excel_path.replace('.xlsx', '_updated.xlsx')

        print(f"Loading Excel file: {excel_path}")
        wb = load_workbook(excel_path)
        ws = wb.active

        df = pd.read_excel(excel_path, header=4)
        print(f"Found {len(df)} rows in Excel")

        pdf_folder_path = Path(pdf_folder)
        pdf_files = list(pdf_folder_path.glob('*/*.pdf'))
        pdf_files += list(pdf_folder_path.glob(f'*/*rop*/*.pdf'))

        if not pdf_files:
            print(f"No PDF files found in {pdf_folder}")
            return

        print(f"Found {len(pdf_files)} PDF files to process\n")

        # Build lookup map from TPN to pdf paths
        pdf_map = {}
        for pf in pdf_files:
            tpn = self.extract_tpn_from_filename(pf.name)
            if tpn:
                pdf_map.setdefault(str(tpn), []).append(pf)
            else:
                pdf_map.setdefault(pf.name, []).append(pf)
        # print(*pdf_map.items())
        if 'TPN No.' not in df.columns:
            print("  ⚠ 'TPN No.' column not found in Excel. Aborting.")
            return
        
        for row_idx, row in df.iterrows():
            print()
            excel_row = row_idx + 6
            # print(pd.isna(row['c) Other fundings if any?']), row['c) Other fundings if any?'])
            if not pd.isna(row.get('Has the institution specified minimum of two faculty members with relevant expertise in quantum technologies, each dedicating at least one-third of their academic time to the lab and its activities? ')):
                
                print(f"Row {excel_row}: already filled, skipping.")
                continue
            tpn_val = row.get('TPN No.')
            if pd.isna(tpn_val):
                print(f"Row {excel_row}: Empty TPN, skipping.")
                continue

            if isinstance(tpn_val, float) and tpn_val.is_integer():
                tpn_str = str(int(tpn_val))
            else:
                tpn_str = str(tpn_val).strip()

            matching_pdfs = pdf_map.get(tpn_str)
            if not matching_pdfs:
                matches = [pf for pf in pdf_files if tpn_str in pf.name]
                if matches:
                    matching_pdfs = matches

            if not matching_pdfs:
                print(f"  ⚠ Could not find PDF for TPN {tpn_str} (Excel row {excel_row})")
                continue

            
            pdf_file = matching_pdfs[0]
            print(f"Processing Excel row {excel_row} - TPN {tpn_str}: {pdf_file.name}")

            extracted_data = self.extract_from_pdf(pdf_file)
            if extracted_data:
                # Update columns L..Z
                ws[f'L{excel_row}'] = extracted_data.get('col_l', '-')
                ws[f'M{excel_row}'] = extracted_data.get('col_m', '-')
                ws[f'N{excel_row}'] = extracted_data.get('col_n', '-')
                ws[f'O{excel_row}'] = extracted_data.get('col_o', '-')
                ws[f'P{excel_row}'] = extracted_data.get('col_p', '-')
                ws[f'Q{excel_row}'] = extracted_data.get('col_q', '-')
                ws[f'R{excel_row}'] = extracted_data.get('col_r', '-')
                ws[f'S{excel_row}'] = extracted_data.get('col_s', '-')
                ws[f'T{excel_row}'] = extracted_data.get('col_t', '-')
                ws[f'U{excel_row}'] = extracted_data.get('col_u', '-')
                ws[f'V{excel_row}'] = extracted_data.get('col_v', '-')
                ws[f'W{excel_row}'] = extracted_data.get('col_w', '-')
                ws[f'X{excel_row}'] = extracted_data.get('col_x', '-')
                ws[f'Y{excel_row}'] = extracted_data.get('col_y', '-')
                ws[f'Z{excel_row}'] = extracted_data.get('col_z', '-')
                ws[f'AA{excel_row}'] = extracted_data.get('col_aa', '-')
                ws[f'AB{excel_row}'] = extracted_data.get('col_ab', '-')


                print(f"  ✓ Updated row {excel_row} in Excel, TPN: {tpn_str}")
            else:
                print(f"  ⚠ Could not extract data from {pdf_file.name}")
            

            wb.save(output_path)
            print()

        print(f"\n{'='*60}")
        print(f"✓ Excel file updated and saved to: {output_path}")
        print(f"{'='*60}")
    
    def extract_tpn_from_filename(self, filename):
        """Extract TPN number from filename if present"""
        import re
        
        # Try to find just numbers
        match = re.search(r'(?<=ProposalID_)\d+(?=_finalproposal)', filename)
        if match:
            return match.group()
        
        return None
    
    def find_row_by_tpn(self, df, tpn, filename):
        """Find row index by TPN number or filename
        
        Args:
            df: DataFrame with Excel data
            tpn: TPN number to search for
            
        Returns:
            Row index (0-based) or None if not found
        """
        # Try to find by TPN No. column
        if 'TPN No.' in df.columns and tpn:
            tpn_col = df['TPN No.'].astype(str)
            matches = tpn_col.str.contains(str(tpn), na=False, case=False)
            if matches.any():
                return matches.idxmax()
        
        return None


# Example usage
if __name__ == "__main__":
    # SETUP: Replace with your configuration
    API_KEY = "AIzaSyCK3xImjoMGO2pIU9g5XXds0TkwyIgpihc"
    
    # Path to your existing Excel file (with headers in row 5)
    EXCEL_FILE = "./proposals_sheet.xlsx"  # Change to your Excel file path
    
    # Path to folder containing PDF proposals
    PDF_FOLDER = "./proposals"  # Change to your PDF folder path
    
    # Output file path (will create new file, won't overwrite original)
    OUTPUT_FILE = "./proposals_sheet_new.xlsx"
    
    # Use Pro model for highest accuracy
    USE_PRO_MODEL = True
    
    print("="*60)
    print("PDF Proposal Extractor - Excel Updater")
    print("="*60)
    print()
    
    # Create extractor instance
    extractor = ProposalExtractorGemini(api_key=API_KEY, use_pro_model=USE_PRO_MODEL)
    
    # Update Excel with extracted data from PDFs
    extractor.update_excel_with_proposals(
        excel_path=EXCEL_FILE,
        pdf_folder=PDF_FOLDER,
        output_path=OUTPUT_FILE
    )