import os
import pandas as pd
import google.generativeai as genai
from pathlib import Path
import json
from openpyxl import load_workbook

class ProposalExtractorGemini:
    def __init__(self, api_key, use_pro_model=True):
        """Initialize with Gemini API key
        
        Args:
            api_key: Your Gemini API key
            use_pro_model: If True, uses gemini-1.5-pro (highest accuracy, slower)
                          If False, uses gemini-1.5-flash (fast, good accuracy)
        """
        genai.configure(api_key=api_key)
        model_name = 'gemini-2.5-pro' if use_pro_model else 'gemini-2.5-flash'
        self.model = genai.GenerativeModel(model_name)
        print(f"Using model: {model_name}")
    
    def extract_from_pdf(self, pdf_path, max_retries=2):
        """Extract data from PDF using Gemini API
        
        Args:
            pdf_path: Path to PDF file
            max_retries: Number of retry attempts if extraction fails
            
        Returns:
            Dictionary with extracted data or None if extraction fails
        """
        for attempt in range(max_retries + 1):
            try:
                # Upload PDF to Gemini
                pdf_file = genai.upload_file(pdf_path)
                
                # Create prompt for extraction
                prompt = """
                Extract the following information from this proposal PDF and return as JSON:
                {
                    "col_l": "value",
                    "col_m": "value",
                    ...
                    "col_z": "value"
                }
                """
                
                # Generate response
                response = self.model.generate_content([pdf_file, prompt])
                response_text = response.text
                
                # Clean up response text
                if response_text.startswith('```json'):
                    response_text = response_text[7:]
                elif response_text.startswith('```'):
                    response_text = response_text[3:]
                
                if response_text.endswith('```'):
                    response_text = response_text.rsplit('```', 1)[0]
                
                response_text = response_text.strip()
                
                # Parse JSON
                data = json.loads(response_text)
                
                # Validate the response has all required fields
                required_fields = [
                    'col_l', 'col_m', 'col_n', 'col_o', 'col_p', 'col_q',
                    'col_r', 'col_s', 'col_t', 'col_u', 'col_v', 'col_w',
                    'col_x', 'col_y', 'col_z'
                ]
                
                missing_fields = [f for f in required_fields if f not in data]
                
                if missing_fields and attempt < max_retries:
                    print(f"  ⚠ Missing fields: {missing_fields}. Retrying (attempt {attempt + 2}/{max_retries + 1})...")
                    continue
                
                # Clean up: delete the uploaded file from Gemini
                pdf_file.delete()
                
                return data
            
            except json.JSONDecodeError as e:
                print(f"  ⚠ Error parsing JSON response: {str(e)}")
                print(f"  Raw response: {response_text[:200]}...")
                return None
            except Exception as e:
                print(f"  ⚠ Error processing PDF: {str(e)}")
                return None
        
        return None
    
    def update_excel_with_proposals(self, excel_path, pdf_folder, output_path=None):
        """Update existing Excel file with extracted data from PDFs
        
        Args:
            excel_path: Path to existing Excel file with headers in row 5
            pdf_folder: Folder containing PDF proposals
            output_path: Path for output file (if None, overwrites original)
        """
        if output_path is None:
            output_path = excel_path.replace('.xlsx', '_updated.xlsx')
        
        # Load the existing Excel file
        print(f"Loading Excel file: {excel_path}")
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Read existing data to find TPN numbers and match with PDFs
        df = pd.read_excel(excel_path, header=4)  # Row 5 is index 4
        
        print(f"Found {len(df)} rows in Excel")
        print(f"Columns: {list(df.columns)[:5]}...")  # Show first 5 columns
        
        # Get list of PDF files (expecting structure like proposals/<TPN>/ProposalID_<TPN>_finalproposal.pdf)
        pdf_folder_path = Path(pdf_folder)
        pdf_files = list(pdf_folder_path.glob('*/*.pdf'))

        if not pdf_files:
            print(f"No PDF files found in {pdf_folder}")
            return

        print(f"Found {len(pdf_files)} PDF files to process\n")

        # Build a mapping from TPN string -> list of PDF Path objects for quick lookup
        pdf_map = {}
        for pf in pdf_files:
            tpn = self.extract_tpn_from_filename(pf.name)
            if tpn:
                pdf_map.setdefault(str(tpn), []).append(pf)
            else:
                # if unable to extract via filename pattern, still keep under full name as fallback
                pdf_map.setdefault(pf.name, []).append(pf)

        # Iterate Excel rows and use the TPN from the sheet to find matching PDF(s)
        for row_idx, row in df.iterrows():
            # Row in Excel = row_idx + 6 (header is row 5, data starts row 6)
            excel_row = row_idx + 6

            if 'TPN No.' not in df.columns:
                print(f"  ⚠ 'TPN No.' column not found in Excel. Aborting.")
                break

            tpn_val = row.get('TPN No.')
            if pd.isna(tpn_val):
                print(f"Row {excel_row}: Empty TPN, skipping.")
                continue

            # Normalize TPN to string (handle floats like 135236.0)
            if isinstance(tpn_val, float) and tpn_val.is_integer():
                tpn_str = str(int(tpn_val))
            else:
                tpn_str = str(tpn_val).strip()

            # Try direct mapping lookup first
            matching_pdfs = pdf_map.get(tpn_str)

            # Fallback: search filenames that contain the TPN string
            if not matching_pdfs:
                matches = [pf for pf in pdf_files if tpn_str in pf.name]
                if matches:
                    matching_pdfs = matches

            if not matching_pdfs:
                print(f"  ⚠ Could not find PDF for TPN {tpn_str} (Excel row {excel_row})")
                continue

            # If multiple PDFs found, pick the first and warn
            if len(matching_pdfs) > 1:
                print(f"  ⚠ Multiple PDFs found for TPN {tpn_str}, using first: {matching_pdfs[0].name}")

            pdf_file = matching_pdfs[0]
            print(f"Processing Excel row {excel_row} - TPN {tpn_str}: {pdf_file.name}")

            # Extract data from PDF
            extracted_data = self.extract_from_pdf(pdf_file)

            if extracted_data:
                # Update columns L through Z
                ws[f'L{excel_row}'] = extracted_data.get('col_l', '-')
                ws[f'M{excel_row}'] = extracted_data.get('col_m', '-')
                ws[f'N{excel_row}'] = extracted_data.get('col_n', '-')
                ws[f'O{excel_row}'] = extracted_data.get('col_o', '-')
                ws[f'P{excel_row}'] = extracted_data.get('col_p', '-')
                ws[f'Q{excel_row}'] = extracted_data.get('col_q', '-')
                ws[f'R{excel_row}'] = extracted_data.get('col_r', '-')
                ws[f'S{excel_row}'] = extracted_data.get('col_s', '-')
                ws[f'T{excel_row}'] = extracted_data.get('col_t', '-')
                ws[f'U{excel_row}'] = extracted_data.get('col_u', '-')
                ws[f'V{excel_row}'] = extracted_data.get('col_v', '-')
                ws[f'W{excel_row}'] = extracted_data.get('col_w', '-')
                ws[f'X{excel_row}'] = extracted_data.get('col_x', '-')
                ws[f'Y{excel_row}'] = extracted_data.get('col_y', '-')
                ws[f'Z{excel_row}'] = extracted_data.get('col_z', '-')

                print(f"  ✓ Updated row {excel_row} in Excel, TPN: {tpn_str}")
            else:
                print(f"  ⚠ Could not extract data from {pdf_file.name}")

            print()

        # Save the updated Excel file after processing all rows
        wb.save(output_path)
        print(f"\n{'='*60}")
        print(f"✓ Excel file updated and saved to: {output_path}")
        print(f"{'='*60}")
    
    def extract_tpn_from_filename(self, filename):
        """Extract TPN number from filename if present"""
        import re
        
        # Try to find just numbers after ProposalID_
        match = re.search(r'(?<=ProposalID_)\d+(?=_finalproposal)', filename)
        if match:
            return match.group()
        
        # Fallback: try to find any sequence of digits
        match = re.search(r'\d+', filename)
        if match:
            return match.group()
        
        return None


# Example usage
if __name__ == "__main__":
    # SETUP: Replace with your configuration
    API_KEY = "AIzaSyCK3xImjoMGO2pIU9g5XXds0TkwyIgpihc"
    
    # Path to your existing Excel file (with headers in row 5)
    EXCEL_FILE = "./proposals_sheet.xlsx"  # Change to your Excel file path
    
    # Path to folder containing PDF proposals
    PDF_FOLDER = "./proposals"  # Change to your PDF folder path
    
    # Output file path (will create new file, won't overwrite original)
    OUTPUT_FILE = "./proposals_sheet_updated.xlsx"
    
    # Use Pro model for highest accuracy
    USE_PRO_MODEL = True
    
    print("="*60)
    print("PDF Proposal Extractor - Excel Updater")
    print("="*60)
    print()
    
    # Create extractor instance
    extractor = ProposalExtractorGemini(api_key=API_KEY, use_pro_model=USE_PRO_MODEL)
    
    # Update Excel with extracted data from PDFs
    extractor.update_excel_with_proposals(
        excel_path=EXCEL_FILE,
        pdf_folder=PDF_FOLDER,
        output_path=OUTPUT_FILE
    )